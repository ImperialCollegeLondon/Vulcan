from openpyxl import load_workbook
import numpy as np
from scipy.stats import qmc
import subprocess

def update_cell(sheet, row, col, value):
        sheet.cell(row=row, column=col).value = str(value)

def read_parameters(parameter_map, filename):
    print(f"Started reading file: {filename}")

    with open(filename, 'r') as file:
        for line in file:
            columns = line.strip().split(',')
            var_name = columns[0]
            var_value = float(columns[1])
            parameter_map[var_name] = var_value

    print(f"Finished reading and closed file: {filename}")


initialization_parameters = {}
main_loop_parameters = {}
randomness_parameters = {}


read_parameters(initialization_parameters, '../InitializationData/Initialization_Parameters_init.csv')
read_parameters(main_loop_parameters, '../InitializationData/Main_Loop_Parameters_init.csv')
read_parameters(randomness_parameters, '../InitializationData/Randomness_Parameters_init.csv')


# Initialization Parameters
household_init_c_f_mean = initialization_parameters['household_init_c_f_mean']
household_init_c_f_min =  initialization_parameters['household_init_c_f_min']
household_init_c_f_max =  initialization_parameters['household_init_c_f_max']

household_init_s_optimist_mean = initialization_parameters['household_init_s_optimist_mean']
household_init_s_optimist_min = initialization_parameters['household_init_s_optimist_min']
household_init_s_optimist_max = initialization_parameters['household_init_s_optimist_max']

household_init_unemp_tolerance_mean = initialization_parameters['household_init_unemp_tolerance_mean']
household_init_unemp_tolerance_min = initialization_parameters["household_init_unemp_tolerance_min"]
household_init_unemp_tolerance_max = initialization_parameters["household_init_unemp_tolerance_max"]

household_init_wealth_mean = initialization_parameters['household_init_wealth_mean']
household_init_wealth_min = 100
household_init_wealth_max = 1250

household_init_minimum_wage_mean = initialization_parameters['household_init_minimum_wage']
household_init_minimum_wage_min = 0
household_init_minimum_wage_max = 900

# Main Loop Parameters


n_households = main_loop_parameters['n_households']
n_households_min = 800
n_households_max = 2500

n_consumer_firms = main_loop_parameters['n_households']
n_consumer_firms_min = 120
n_consumer_firms_max = 480

n_capital_firms = main_loop_parameters['n_households']
n_capital_firms_min = 40
n_capital_firms_max = 120

household_targeted_savings_to_income_ratio = main_loop_parameters['household_targeted_savings_to_income_ratio']
household_targeted_savings_to_income_ratio_min = 2
household_targeted_savings_to_income_ratio_max = 5

firm_cons_inv_reaction_factor_mean = main_loop_parameters['firm_cons_inv_reaction_factor']
firm_cons_inv_reaction_factor_min = 0.99
firm_cons_inv_reaction_factor_max = 1

# Randomness Parameters
firm_cons_rand_desired_inventory_factor_change = randomness_parameters["firm_cons_rand_desired_inventory_factor_change"]
firm_cons_rand_desired_inventory_factor_change_min = 0
firm_cons_rand_desired_inventory_factor_change_max = 0.1

firm_cons_rand_price_change_upper_limit = randomness_parameters["firm_cons_rand_price_change_upper_limit"]
firm_cons_rand_price_change_upper_limit_min = 0
firm_cons_rand_price_change_upper_limit_max = 1

firm_cons_rand_wage_change = randomness_parameters["firm_cons_rand_wage_change"]
firm_cons_rand_wage_change_min = 0
firm_cons_rand_wage_change_max = 1

firm_cap_rand_desired_inventory_factor_change = randomness_parameters["firm_cap_rand_desired_inventory_factor_change"]
firm_cap_rand_desired_inventory_factor_change_min = 0
firm_cap_rand_desired_inventory_factor_change_max = 1

firm_cap_rand_wage_change = randomness_parameters["firm_cap_rand_wage_change"]
firm_cap_rand_wage_change_min = 0
firm_cap_rand_wage_change_max = 1

def init_params():
    sample = np.array([household_init_c_f_mean,
        household_init_s_optimist_mean,
        household_init_wealth_mean,
        household_init_unemp_tolerance_mean,
        household_init_minimum_wage_mean,
        
        n_households,
        n_consumer_firms,
        n_capital_firms,
        household_targeted_savings_to_income_ratio,
        firm_cons_inv_reaction_factor_mean,
        
        firm_cons_rand_desired_inventory_factor_change,
        firm_cons_rand_price_change_upper_limit,
        firm_cons_rand_wage_change,
        firm_cap_rand_desired_inventory_factor_change,
        firm_cap_rand_wage_change])
    
    return sample
    
    
def change_parameters():

    # Load workbook and sheets
    wb = load_workbook(filename='../InitializationData/Simulation_Parameters_main.xlsm', read_only=False, keep_vba=True)
    Initialization_Parameters_wb = wb['Initialization_Parameters']
    Main_Loop_Parameters_wb = wb['Main_Loop_Parameters']
    Randomness_Parameters_wb = wb['Randomness_Parameters']

    # Define the parameter bounds
    param_bounds = [
        (household_init_c_f_min, household_init_c_f_max),
        (household_init_s_optimist_min, household_init_s_optimist_max),
        (household_init_wealth_min, household_init_wealth_max),
        (household_init_unemp_tolerance_min, household_init_unemp_tolerance_max),
        (household_init_minimum_wage_min, household_init_minimum_wage_max),
        
        (n_households_min, n_households_max),
        (n_consumer_firms_min, n_consumer_firms_max),
        (n_capital_firms_min, n_capital_firms_max),
        (household_targeted_savings_to_income_ratio_min, household_targeted_savings_to_income_ratio_max),
        (firm_cons_inv_reaction_factor_min, firm_cons_inv_reaction_factor_max),
        
        (firm_cons_rand_desired_inventory_factor_change_min, firm_cons_rand_desired_inventory_factor_change_max),
        (firm_cons_rand_price_change_upper_limit_min, firm_cons_rand_price_change_upper_limit_max),
        (firm_cons_rand_wage_change_min, firm_cons_rand_wage_change_max),
        (firm_cap_rand_desired_inventory_factor_change_min, firm_cap_rand_desired_inventory_factor_change_max),
        (firm_cap_rand_wage_change_min, firm_cap_rand_wage_change_max)
    ]

    num_params = len(param_bounds)
    num_samples = 1 

    sampler = qmc.LatinHypercube(d=num_params)
    sample = sampler.random(n=num_samples)
    sample = qmc.scale(sample, [b[0] for b in param_bounds], [b[1] for b in param_bounds])
    sample = np.round(sample, 2)
    sample[0,5] = int(sample[0, 5])
    sample[0,6] = int(sample[0, 6])
    sample[0,7] = int(sample[0, 7])
    sample_str = np.array([str(x).replace('.', ',') for x in sample[0]])
    
    row_indices_Initialization_Parameters = [5, 13, 21, 25, 34]
    row_indices_Main_Loop_Parameters = [4, 5, 6, 11, 19]
    row_indices_Randomness_Parameters = [10, 11, 12, 16, 18]
    col_index = 2  

    for i, row in enumerate(row_indices_Initialization_Parameters):
        update_cell(Initialization_Parameters_wb, row, col_index, sample_str[i])
    for i, row in enumerate(row_indices_Main_Loop_Parameters):
        update_cell(Main_Loop_Parameters_wb, row, col_index, sample_str[i+len(row_indices_Initialization_Parameters)])
    for i, row in enumerate(row_indices_Randomness_Parameters):
        update_cell(Randomness_Parameters_wb, row, col_index, sample_str[i+len(row_indices_Initialization_Parameters)+len(row_indices_Main_Loop_Parameters)])

    wb.save('../InitializationData/Simulation_Parameters_main.xlsm')
    print('Workbook saved to', '../InitializationData/Simulation_Parameters_main.xlsm')
    
    applescript_path = '/Users/ayman/Desktop/FYP_project/Vulcan/script_vba.scpt'
    subprocess.run(['osascript', applescript_path])

    return sample

def set_parameters(Theta):
    sample_str = np.array([str(x).replace('.', ',') for x in Theta])
    
    wb = load_workbook(filename='../InitializationData/Simulation_Parameters_main.xlsm', read_only=False, keep_vba=True)
    Initialization_Parameters_wb = wb['Initialization_Parameters']
    Main_Loop_Parameters_wb = wb['Main_Loop_Parameters']
    Randomness_Parameters_wb = wb['Randomness_Parameters']
    
    row_indices_Initialization_Parameters = [5, 13, 21, 25, 34]
    row_indices_Main_Loop_Parameters = [4, 5, 6, 11, 19]
    row_indices_Randomness_Parameters = [10, 11, 12, 16, 18]
    col_index = 2  
    
    for i, row in enumerate(row_indices_Initialization_Parameters):
        update_cell(Initialization_Parameters_wb, row, col_index, sample_str[i])
    for i, row in enumerate(row_indices_Main_Loop_Parameters):
        update_cell(Main_Loop_Parameters_wb, row, col_index, sample_str[i+len(row_indices_Initialization_Parameters)])
    for i, row in enumerate(row_indices_Randomness_Parameters):
        update_cell(Randomness_Parameters_wb, row, col_index, sample_str[i+len(row_indices_Initialization_Parameters)+len(row_indices_Main_Loop_Parameters)])

    # Save the workbook
    wb.save('../InitializationData/Simulation_Parameters_main.xlsm')
    
    applescript_path = '/Users/ayman/Desktop/FYP_project/Vulcan/script_vba.scpt'
    subprocess.run(['osascript', applescript_path])
        

sample = change_parameters()